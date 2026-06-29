# -*- coding: utf-8 -*-
"""
Extrai o catálogo do .xlsx do Simulador Clover para um JSON (catalogo-seed.json)
que o importador Node (src/cli/importar-catalogo.js) injeta no Postgres.

Reaproveita o parsing validado com openpyxl. Rode com:
    python server/scripts/extrair_catalogo.py "<caminho do .xlsx>" "<saida.json>"
"""
import sys, os, json, datetime
import openpyxl

def t2min(v):
    """Converte tempo Excel -> minutos inteiros. Aceita timedelta/time/datetime/number/str."""
    if v is None:
        return 0
    if isinstance(v, datetime.timedelta):
        return round(v.total_seconds() / 60)
    if isinstance(v, datetime.time):
        return v.hour * 60 + v.minute + round(v.second / 60)
    if isinstance(v, datetime.datetime):
        return v.hour * 60 + v.minute + round(v.second / 60)
    if isinstance(v, (int, float)):
        # fração de dia (0.5 = 12h) — heurística: <10 provavelmente é fração
        return round(float(v) * 24 * 60)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return 0
        parts = s.split(":")
        try:
            if len(parts) == 3:
                h, m, sec = (int(p) for p in parts)
                return h * 60 + m + round(sec / 60)
            if len(parts) == 2:
                h, m = (int(p) for p in parts)
                return h * 60 + m
        except ValueError:
            return 0
    return 0

def norm(s):
    return None if s is None else str(s).strip()

def extrair(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Funcionalidades -> modulos + funcionalidades ----------------------
    ws = wb["Funcionalidades"]
    # Linhas que iniciam um módulo (col B não-vazia), exceto o cabeçalho R2.
    bordas = []
    for r in range(3, ws.max_row + 1):
        b = norm(ws.cell(row=r, column=2).value)
        if b:
            bordas.append((r, b))

    modulos = []
    funcionalidades = []
    for idx, (r_ini, nome_mod) in enumerate(bordas):
        r_fim = (bordas[idx + 1][0] - 1) if idx + 1 < len(bordas) else ws.max_row
        modulos.append({"nome": nome_mod, "ordem": idx})
        mae_atual = None
        for r in range(r_ini, r_fim + 1):
            mae_cel = norm(ws.cell(row=r, column=3).value)  # Funcionalidade Mãe
            if mae_cel:
                mae_atual = mae_cel
            filha = norm(ws.cell(row=r, column=5).value)     # Funcionalidade Filha
            if not filha:
                continue
            tipo = norm(ws.cell(row=r, column=6).value)
            horas = t2min(ws.cell(row=r, column=7).value)
            funcionalidades.append({
                "modulo": nome_mod,
                "mae": mae_atual,
                "nome": filha,
                "tipo": tipo,
                "horas_minutos": horas,
                # pacote padrão = obrigatórias pré-marcadas (ajustável depois no app)
                "pacote_padrao": bool(tipo and tipo.lower().startswith("obrigat")),
                "ordem": r,
            })

    # ---- ERPs e métodos de integração (aba Valor) --------------------------
    wv = wb["Valor"]
    erps = []
    for i, r in enumerate(range(2, 23)):
        a = norm(wv.cell(row=r, column=1).value)
        if a:
            erps.append({"nome": a, "ordem": i})
    metodos = []
    for i, r in enumerate(range(2, 8)):
        d = norm(wv.cell(row=r, column=4).value)
        if d and not d.startswith("="):
            metodos.append({"nome": d, "ordem": i})

    # ---- Fluxos de integração (Solution / ERP Terceiro) --------------------
    # Matriz de ativação: colunas H..AB (8..28) da linha 2 são os MÓDULOS; nas
    # linhas de fluxo, a célula "SIM" indica que aquele fluxo é ativado por aquele
    # módulo. Capturamos a lista de módulos que ativam cada fluxo.
    MOD_COL_INI, MOD_COL_FIM = 8, 28  # H..AB
    fluxos = []
    abas = {
        "solution": "Aliare Integra Solution",
        "erp_terceiro": "Aliare Integra ERP Terceiro",
    }
    for contexto, aba in abas.items():
        if aba not in wb.sheetnames:
            continue
        wsi = wb[aba]
        # nomes dos módulos no cabeçalho (linha 2)
        modulos_cols = {}
        for c in range(MOD_COL_INI, MOD_COL_FIM + 1):
            nome_mod = norm(wsi.cell(row=2, column=c).value)
            if nome_mod and nome_mod.upper().startswith("CRM"):
                modulos_cols[c] = nome_mod
        grupo = None
        for r in range(3, wsi.max_row + 1):
            a = norm(wsi.cell(row=r, column=1).value)  # Fluxos (grupo)
            if a:
                grupo = a
            tabela = norm(wsi.cell(row=r, column=2).value)
            tipo = norm(wsi.cell(row=r, column=3).value)
            if not tabela and not a:
                continue
            ativa = []
            for c, nome_mod in modulos_cols.items():
                v = norm(wsi.cell(row=r, column=c).value)
                if v and v.strip().upper() == "SIM":
                    ativa.append(nome_mod)
            fluxos.append({
                "contexto": contexto,
                "fluxo": grupo,
                "tabela": tabela,
                "tipo": tipo,
                "min_config": t2min(wsi.cell(row=r, column=4).value),
                "min_teste_carga": t2min(wsi.cell(row=r, column=5).value),
                "min_apoio": t2min(wsi.cell(row=r, column=6).value),
                "min_validacao": t2min(wsi.cell(row=r, column=7).value),
                "modulos_ativa": ativa,
                "ordem": r,
            })

    return {
        "modulos": modulos,
        "funcionalidades": funcionalidades,
        "erps": erps,
        "metodos_integracao": metodos,
        "fluxos_integracao": fluxos,
    }

def main():
    path = sys.argv[1] if len(sys.argv) > 1 else r"C:\projetos\PROJETOS IA\Simulador Clover MÁQUINAS - Por funcionalidade - v11052026.1 (1).xlsx"
    saida = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "catalogo-seed.json")
    dados = extrair(path)
    with open(saida, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    print(f"OK -> {saida}")
    print(f"  modulos={len(dados['modulos'])} funcionalidades={len(dados['funcionalidades'])} "
          f"erps={len(dados['erps'])} metodos={len(dados['metodos_integracao'])} fluxos={len(dados['fluxos_integracao'])}")
    tot = sum(f['horas_minutos'] for f in dados['funcionalidades'])
    print(f"  soma horas catalogo = {tot} min = {tot/60:.2f} h")

if __name__ == "__main__":
    main()
